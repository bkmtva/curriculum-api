import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
import pandas as pd
import json

SOURCE_PATH = "/app/excel_files/"
BASE_FILE = SOURCE_PATH+"template.xlsx"
file_save = "curriculum.xlsx"

# response ="C:/Users/anarb/diploma/response.json"
# with open(response, encoding='utf8') as response_file:
#     dict_data = json.load(response_file)

# file_path = "C:/Users/anarb/diploma/template.xlsx"
# saved_file_path = "C:/Users/anarb/diploma/curriculum.xlsx"


def get_dataframe(dict_data):
    try:
        # Check if dict_data is None or does not contain 'courses'
        if not dict_data or 'courses' not in dict_data:
            return pd.DataFrame()  # Return empty DataFrame
        
        # Extract 'courses' from the dictionary
        all_courses = dict_data.get('courses', [])
        
        # Check if 'courses' is empty
        if not all_courses:
            return pd.DataFrame()  # Return empty DataFrame
        
        formatted_data = []

        # Process courses
        for course_data in all_courses:
            course_info = course_data.get('course')
            if not course_info:
                continue  # Skip if there's no course info
            
            subcourses = course_info.get('subcourses', [])
            
            # If there are subcourses, iterate over them
            if subcourses:
                for subcourse in subcourses:
                    formatted_data.append({
                        'semester': course_data['semester'],
                        'course_id': course_data['course_id'],
                        'order_in_semester': course_data['order_in_semester'],
                        'title': subcourse['title'],
                        'title_kz': subcourse.get('title_kz', course_info['title_kz']), 
                        'title_ru': subcourse.get('title_ru', course_info['title_ru']), 
                        'teor': subcourse['teor'],
                        'cr': subcourse['cr'],
                        'term': subcourse['term'],
                        'course_code': subcourse['course_code'],
                        'pr': subcourse['pr'],
                        'ects': subcourse['ects'],
                        'subcourse_title': None  # Set to None for subcourses
                    })
            else:
                # If there are no subcourses, add the main course with None as subcourse_title
                formatted_data.append({
                    'semester': course_data['semester'],
                    'course_id': course_data['course_id'],
                    'order_in_semester': course_data['order_in_semester'],
                    'title': course_info['title'],
                    'title_kz': course_info['title_kz'],
                    'title_ru': course_info['title_ru'],
                    'teor': course_info['teor'],
                    'cr': course_info['cr'],
                    'term': course_info['term'],
                    'course_code': course_info['course_code'],
                    'pr': course_info['pr'],
                    'ects': course_info['ects'],
                    'subcourse_title': None  # Set to None for main courses without subcourses
                })

        # Create DataFrame and additional columns
        df = pd.DataFrame(formatted_data)
        df['combined_titles'] = df['title_kz'] + ' / ' + df['title_ru'] + ' / ' + df['title']
        df['pr_value'] = df['pr'].apply(lambda x: sum(map(int, x.split('+'))))
        df['cr'] = df['cr'].astype(int)
        df['ects'] = df['ects'].astype(int)

        return df

    except Exception:
        # If any error occurs, return an empty DataFrame
        return pd.DataFrame()


def to_excel(curriculum: str, file_save: str ="curriculum.xlsx"):
    dict_data = json.loads(curriculum)
    df = get_dataframe(dict_data)
    if df.empty:
        return BASE_FILE
    
    # Load the Excel file
    wb = load_workbook(filename=BASE_FILE)
    ws = wb['Curriculum']

    # Starting column and row to write the data
    total_cols = 9
    start_col = 4
    start_row = 27

    # Filter the DataFrame by semester and create a list of unique semesters
    semesters = df['semester'].unique()

    # Sort semesters in ascending order
    semesters.sort()

    # Initialize font properties
    bold_font = Font(name='Times New Roman', size=12, bold=True)
    regular_font = Font(name='Times New Roman', size=12)
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    last_row = 27
    # Writing the data for each semester
    for semester in semesters:
        # Filter DataFrame by semester
        df_semester = df[df['semester'] == semester]

        # Sort courses by order_in_semester
        df_semester = df_semester.sort_values(by='order_in_semester')

        # Merge cells for semester header
        merge_range = f'{get_column_letter(1)}{start_row}:{get_column_letter(total_cols)}{start_row}'
        ws.merge_cells(merge_range)

        # Write semester header with bold font
        ws[f'{get_column_letter(1)}{start_row}'].value = f"{semester} семестр"
        ws[f'{get_column_letter(1)}{start_row}'].alignment = Alignment(horizontal='center')
        ws[f'{get_column_letter(1)}{start_row}'].font = bold_font

        # Increment the start_row
        start_row += 1

        # Merge cells for CR total
        cr_start_row = start_row

        # Merge cells for ECTS total
        ects_start_row = start_row

        # Initialize variables to track previous order_in_semester and max CR/ECTS
        prev_order = None
        max_cr = 0
        max_ects = 0

        # Write data for each course in the semester
        for r, course in enumerate(df_semester.itertuples(), start=start_row):
            # If it's not the first row and the order_in_semester is the same as the previous one
            if prev_order == course.order_in_semester and r != start_row:
                # Merge cells for course_code and set the value to None
                ws.merge_cells(f'{get_column_letter(start_col)}{r}:{get_column_letter(start_col)}{r}')
               # ws[f'{get_column_letter(start_col)}{r}'].value = None
            #else:
                # Write course_code
                #ws[f'{get_column_letter(start_col)}{r}'].value = course.course_code

            # Write other columns
            ws[f'{get_column_letter(start_col)}{r}'].value = course.course_code
            ws[f'{get_column_letter(start_col + 1)}{r}'].value = course.combined_titles
            ws[f'{get_column_letter(start_col + 2)}{r}'].value = course.cr
            ws[f'{get_column_letter(start_col + 3)}{r}'].value = course.ects
            ws[f'{get_column_letter(start_col + 4)}{r}'].value = course.pr_value

            # Track CR and ECTS for merging
            if prev_order == course.order_in_semester:
                max_cr = max(max_cr, course.cr)
                max_ects = max(max_ects, course.ects)
            else:
                if prev_order is not None:
                    # Merge cells for CR with max value
                    merge_range_cr = f'{get_column_letter(start_col + 2)}{cr_start_row}:{get_column_letter(start_col + 2)}{r - 1}'
                    ws.merge_cells(merge_range_cr)
                    ws[f'{get_column_letter(start_col + 2)}{cr_start_row}'].value = max_cr

                    # Merge cells for ECTS with max value
                    merge_range_ects = f'{get_column_letter(start_col + 3)}{ects_start_row}:{get_column_letter(start_col + 3)}{r - 1}'
                    ws.merge_cells(merge_range_ects)
                    ws[f'{get_column_letter(start_col + 3)}{ects_start_row}'].value = max_ects

                # Update variables for the new group of courses
                cr_start_row = r
                ects_start_row = r
                max_cr = course.cr
                max_ects = course.ects
                prev_order = course.order_in_semester

            # Set font for each cell
            for col in range(start_col, start_col + 5):
                ws[f'{get_column_letter(col)}{r}'].font = regular_font
              #  ws[f'{get_column_letter(col)}{r}'].border = thin_border

        # Merge cells for CR and ECTS with max value for the last group of courses
        merge_range_cr = f'{get_column_letter(start_col + 2)}{cr_start_row}:{get_column_letter(start_col + 2)}{start_row + len(df_semester) - 1}'
        ws.merge_cells(merge_range_cr)
        ws[f'{get_column_letter(start_col + 2)}{cr_start_row}'].value = max_cr

        merge_range_ects = f'{get_column_letter(start_col + 3)}{ects_start_row}:{get_column_letter(start_col + 3)}{start_row + len(df_semester) - 1}'
        ws.merge_cells(merge_range_ects)
        ws[f'{get_column_letter(start_col + 3)}{ects_start_row}'].value = max_ects

        # Set wrap text and center alignment for the combined_titles column
        for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(df_semester) - 1, min_col=start_col + 1, max_col=start_col + 1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center')

        # Write 'Барлығы/Итого/Total' at the end of each semester
        ws[f'{get_column_letter(start_col)}{start_row + len(df_semester)}'].value = 'Барлығы/Итого/Total'
        ws.merge_cells(f'{get_column_letter(start_col)}{start_row + len(df_semester)}:{get_column_letter(start_col + 1)}{start_row + len(df_semester)}')

        # Write total cr and ects
        df_semester['cr'] = df_semester['cr'].astype(int)
        df_semester['ects'] = df_semester['ects'].astype(int)

        total_cr = sum(df_semester.groupby(['semester','order_in_semester'])['cr'].max())
        total_ects = sum(df_semester.groupby(['semester','order_in_semester'])['ects'].max())
        ws[f'{get_column_letter(start_col + 2)}{start_row + len(df_semester)}'].value = total_cr
        ws[f'{get_column_letter(start_col + 3)}{start_row + len(df_semester)}'].value = total_ects

        # Set font and alignment for total row
        for col in range(start_col, start_col + 4):
            ws[f'{get_column_letter(col)}{start_row + len(df_semester)}'].font = bold_font
            ws[f'{get_column_letter(col)}{start_row + len(df_semester)}'].alignment = Alignment(horizontal='center')


        # Update start_row for the next semester
        start_row += len(df_semester) + 1
        last_row = start_row

    # Calculate total 'cr' and 'ects' for all semesters
    total_cr_all = sum(df.groupby(['semester','order_in_semester'])['cr'].max())
    total_ects_all = sum(df.groupby(['semester','order_in_semester'])['ects'].max())


    # Merge cells for total
    total_merge_range = f'{get_column_letter(start_col)}{last_row}:{get_column_letter(start_col + 1)}{last_row}'
    ws.merge_cells(total_merge_range)
    ws[f'{get_column_letter(start_col)}{last_row}'].value = 'Барлығы/Итого/Total'
    ws[f'{get_column_letter(start_col)}{last_row}'].font = bold_font

    ws[f'{get_column_letter(start_col + 2)}{last_row}'].value = total_cr_all
    ws[f'{get_column_letter(start_col + 3)}{last_row}'].value = total_ects_all
    ws[f'{get_column_letter(start_col + 2)}{last_row}'].font = bold_font
    ws[f'{get_column_letter(start_col + 3)}{last_row}'].font = bold_font
    #ws[f'{get_column_letter(start_col)}{last_row}'].font = Alignment(horizontal='center')


    # Making borders for cells
    for row in range(27, last_row + 1):
        for col in range(1, 10 + 1):
            ws.cell(row=row, column=col).border = thin_border

    # Set program information
    program_code ='6B'+ dict_data.get('program').get('code') 
    program_name =  dict_data.get('program').get('title_kz')
    degree_name =  dict_data.get('degree_name')
    program_info = str(program_code) + ' ' + str(program_name)
    program_in_kz = 'Білім беру бағдарламасы: ' + program_info
    program_in_ru = 'Образовательная программа: ' + program_info
    program_in_en = 'Program: ' + program_info

    degree_info = str(degree_name) + ' ' + str(program_code)
    degree_in_kz = 'Берілетін дәреже: ' + degree_info
    degree_in_ru = 'Присуждаемая степень: '+ degree_info
    degree_in_en = 'Degree: ' + degree_info

    # Set the values in the specified cells
    ws['A17'].value = program_in_kz
    ws['E17'].value = program_in_en
    ws['G17'].value = program_in_ru

    ws['A18'].value = degree_in_kz
    ws['E18'].value = degree_in_en
    ws['G18'].value = degree_in_ru
    
    ws['A17'].font = bold_font
    ws['E17'].font = bold_font
    ws['G17'].font = bold_font
    
    ws['A18'].font = bold_font
    ws['E18'].font = bold_font
    ws['G18'].font = bold_font

    current_year = int(dict_data.get('year'))
    finish_year = current_year
    if degree_name=='B':
        finish_year = finish_year+4
    elif  degree_name=='Master':
        finish_year = finish_year+2
    elif  degree_name=='PhD':
        finish_year = finish_year+3  
        
    years=str(current_year)+'-'+str(finish_year)
    academic_year =years+' оқу жылдарына/ на ' + years+' учебные годы / '+years+ ' academic years'
    ws['A20'].value = academic_year
    ws['A20'].font = bold_font
    
    # Director,Head,Dean adding
    
    merge_range = 'A{0}:G{0}'
    head_dean_dir = ['Оқу-әдістемелік орталық директоры/ Директор УМЦ/ Director of the Educational and Methodical Center ',
                     'Факультет деканы/ Декан факультета/ Dean of the Faculty ',
                     'Кафедра меңгерушісі/ Заведующий кафедрой/ Head of the department ']
    for index, text in enumerate(head_dean_dir):
        row = last_row+1+index
        ws.merge_cells(merge_range.format(row))
        ws[f'A{row}'].value = text +  "________________________"
        ws[f'A{row}'].font = bold_font 
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

    # Save the workbook
    saved_file_path = SOURCE_PATH + file_save
    wb.save(saved_file_path)
    
    return file_save


# print(to_excel(dict_data))